#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import sys
from pathlib import Path
from urllib.parse import urlparse
from typing import Any

from dotenv import load_dotenv

from shanghai_policy_crawler.firecrawl_crawler import FirecrawlPolicyCrawler
from shanghai_policy_crawler.crawl4ai_crawler import Crawl4AIPolicyCrawler
from shanghai_policy_crawler.utils import compact_dict

# Configure Logging
logger = logging.getLogger("crawler")


def load_env_credentials() -> tuple[str | None, str | None, str | None]:
    load_dotenv()
    api_key = os.getenv("FIRECRAWL_API_KEY")
    api_url = os.getenv("FIRECRAWL_API_URL")
    cookie = os.getenv("QIZHIDAO_COOKIE")
    return api_key, api_url, cookie


def build_exporters_from_records(records: list[dict[str, Any]], output_prefix: Path) -> None:
    """
    Compile a polished summary of scraped policies into multiple formats:
    JSON, CSV, Markdown, and Excel (XLSX). Matches Scrapy's summary pipeline for consistency.
    """
    if not records:
        logger.warning("No records collected. Skipping summary compilation.")
        return

    headers = [
        "序号",
        "政策名称",
        "政策分类",
        "政策行业信息（什么类型企业可以申报）",
        "申报条件（满足什么条件可以申请）",
        "补贴形式（前补贴/后补贴/税惠/非补贴）",
        "补贴金额（最高金额/占比/固定额度）",
        "持续时间",
        "补贴申报地点",
        "如何申请补贴",
        "发布/主管部门",
        "地区",
        "企知道链接",
        "政府原贴",
        "正文",
        "数据备注",
    ]

    summary_rows = []
    for index, record in enumerate(records, start=1):
        title = record.get("title") or ""
        content = record.get("content") or ""
        categories = record.get("category_labels") or []
        
        # Pull LLM extracted attributes directly, falling back to text parsing markers if absent
        policy_object = record.get("policy_object") or "见正文"
        policy_conditions = record.get("policy_conditions") or "见正文"
        payment_standard = record.get("payment_standard") or "见正文"
        application_period = record.get("application_period") or record.get("publish_date") or "未明确"
        contact_information = record.get("contact_information") or "见正文"
        payment_method = record.get("payment_method") or "奖励/补贴"
        source = record.get("source") or "未知"
        city = record.get("city") or "上海市"
        
        summary_rows.append({
            "序号": index,
            "政策名称": title,
            "政策分类": "；".join(categories) if isinstance(categories, list) else str(categories),
            "政策行业信息（什么类型企业可以申报）": policy_object,
            "申报条件（满足什么条件可以申请）": policy_conditions,
            "补贴形式（前补贴/后补贴/税惠/非补贴）": payment_method,
            "补贴金额（最高金额/占比/固定额度）": payment_standard,
            "持续时间": application_period,
            "补贴申报地点": contact_information,
            "如何申请补贴": "以政策详情和申报指南为准",
            "发布/主管部门": source,
            "地区": city,
            "企知道链接": record.get("url") or "",
            "政府原贴": record.get("government_original_url") or "见详情",
            "正文": content,
            "数据备注": "Firecrawl LLM 结构化抽取整理；请以官方渠道及政府原贴为准",
        })

    output_prefix.parent.mkdir(parents=True, exist_ok=True)

    # 1. JSON
    json_path = output_prefix.with_suffix(".json")
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(summary_rows, f, ensure_ascii=False, indent=2)
    logger.info("Saved JSON summary: %s", json_path)

    # 2. CSV
    csv_path = output_prefix.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(summary_rows)
    logger.info("Saved CSV summary: %s", csv_path)

    # 3. Markdown (RAG friendly format)
    md_path = output_prefix.with_suffix(".md")
    lines = [
        "# Firecrawl 政策补贴信息汇总",
        "",
        f"- 采集记录数：{len(summary_rows)}",
        "- 说明：由 Firecrawl LLM 智能提取并总结，正式申报时请以官方政府公告与申报入口为准。",
        "",
    ]
    for row in summary_rows:
        lines.append(f"## {row['序号']}. {row['政策名称']}")
        lines.append("")
        for key in headers[2:14]:
            lines.append(f"- **{key}**：{row.get(key) or '未明确'}")
        lines.append("")
    with md_path.open("w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    logger.info("Saved Markdown summary: %s", md_path)

    # 4. Excel (using openpyxl if available)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "政策补贴整理"
        ws.append(headers)

        for row in summary_rows:
            ws.append([row.get(h, "") for h in headers])

        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        widths = [8, 38, 20, 44, 54, 30, 42, 28, 42, 56, 24, 18, 42, 42, 80, 36]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
        xlsx_path = output_prefix.with_suffix(".xlsx")
        wb.save(xlsx_path)
        logger.info("Saved Excel workbook summary: %s", xlsx_path)
    except ImportError:
        logger.warning("openpyxl is not installed. Skipping Excel compilation.")


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )
    parser = argparse.ArgumentParser(description="Shanghai Government Policy Crawler (shanghai.gov.cn / zwdt.sh.gov.cn).")
    parser.add_argument("--start-url", default="https://www.shanghai.gov.cn/zhengce/more?level=city", help="Listing URL to scrape links from")
    parser.add_argument("--max-pages", type=int, default=1, help="Max list pages to check")
    parser.add_argument("--output-file", default="output/firecrawl_data.jsonl", help="Path to write the main JSONL records")
    parser.add_argument("--summary-prefix", default="output/firecrawl_policy_summary", help="Prefix path for RAG md, Excel and CSV formats")
    parser.add_argument("--api-key", default=None, help="Firecrawl API Key (optional override)")
    parser.add_argument("--api-url", default=None, help="Firecrawl Custom API URL (optional override)")
    parser.add_argument("--keyword", default="", help="Keyword filter")
    parser.add_argument("--min-year", type=int, default=2024, help="Exclude policies published before this year")
    parser.add_argument("--city-filter", default="上海市", help="Exclude policies without this city tag")
    parser.add_argument("--cookie", default=None, help="Cookie string for authenticating request headers (optional)")
    parser.add_argument("--engine", default="crawl4ai", choices=["crawl4ai", "firecrawl"], help="Crawling engine (default: crawl4ai)")
    args = parser.parse_args()

    # Credentials loading sequence: CLI Arg -> Env/Dotenv
    api_key, api_url, env_cookie = load_env_credentials()
    final_api_key = args.api_key or api_key
    final_api_url = args.api_url or api_url
    final_cookie = args.cookie or env_cookie

    if args.engine == "firecrawl":
        if not final_api_key:
            logger.critical(
                "Firecrawl API Key is missing! Set FIRECRAWL_API_KEY in your .env or pass it via --api-key argument."
            )
            sys.exit(1)

        logger.info("Initializing Firecrawl Policy Crawler...")
        crawler = FirecrawlPolicyCrawler(
            api_key=final_api_key,
            api_url=final_api_url,
            city_filter=args.city_filter,
            min_year=args.min_year,
            keyword=args.keyword,
            cookie=final_cookie,
        )
    else:
        logger.info("Initializing Crawl4AI Policy Crawler...")
        crawler = Crawl4AIPolicyCrawler(
            city_filter=args.city_filter,
            min_year=args.min_year,
            keyword=args.keyword,
            cookie=final_cookie,
        )

    # CSS selectors for harvesting policy detail links from list pages
    # Covers two platforms:
    #   shanghai.gov.cn  — policy pages use /{section}/{YYYYMMDD}/{32hex}.html pattern
    #   zwdt.sh.gov.cn   — uses /policy/project-detail?id=... pattern
    detail_link_selectors = [
        # shanghai.gov.cn: year-based date directory (2020~2030)
        "a[href*='/202'][href$='.html']",
        "a[href*='/203'][href$='.html']",
        # shanghai.gov.cn: nw-style legacy URLs
        "a[href*='/nw4411/']",
        "a[href*='/nw2314/']",
        # zwdt.sh.gov.cn
        "a[href*='project-detail']",
        "a[href*='/policy/']",
        # fallback legacy patterns
        "a[href*='/docs/']",
        "a[href*='/declare/']",
        "a[href*='/publicity/pub/']",
        "a[href*='article-zcview']",
        "a[href*='/lead/']",
    ]

    # Run scraping sequence
    detail_urls = []
    # Patterns that identify a URL as a direct policy detail page (skip list harvesting)
    # For shanghai.gov.cn, detail URLs contain a date directory like /20230627/
    # Subsite list pages (e.g. /jcsfbzsrx/index.html) do NOT contain date dirs
    detail_patterns = [
        "/docs/", "/declare/", "/publicity/pub/", "article-zcview", "/lead/",
        "project-detail", "businessId=",
        # shanghai.gov.cn date-directory pattern (e.g. /20230327/ or /20251018/)
        "/2020", "/2021", "/2022", "/2023", "/2024", "/2025", "/2026",
    ]
    # API-based list pages (SPA that exposes a REST API)
    api_list_patterns = [
        "/zhengce/more",
        "/zhengce/list",
        "/gwk/policy",
    ]
    if Path(args.start_url).is_file():
        logger.info("Loading URLs from seed file: %s", args.start_url)
        with open(args.start_url, "r", encoding="utf-8") as f:
            seed_lines = [line.strip() for line in f if line.strip() and not line.strip().startswith("#")]
        # Route seed lines: API-list lines → API harvest; detail lines → direct; subsite lines → Firecrawl
        from shanghai_policy_crawler.shanghai_gov_api import ShanghaiGovApiClient, DETAIL_URL_TPL
        for seed_url in seed_lines:
            parsed = urlparse(seed_url)
            if any(pat in seed_url for pat in api_list_patterns):
                # Determine level from query string
                from urllib.parse import parse_qs
                qs = parse_qs(parsed.query)
                level = qs.get("level", ["city"])[0]
                site_id_map = {"city": ["0001"], "department": None, "district": None}
                site_ids = site_id_map.get(level, ["0001"])
                logger.info("API harvest for level=%s, site_ids=%s", level, site_ids)
                api_client = ShanghaiGovApiClient(site_id_list=site_ids, page_size=20)
                for rec in api_client.iter_records(max_pages=args.max_pages, min_year=args.min_year):
                    url = rec["detail_url"]
                    if url not in detail_urls:
                        detail_urls.append(url)
            elif any(pat in seed_url for pat in detail_patterns):
                if seed_url not in detail_urls:
                    detail_urls.append(seed_url)
            else:
                # Firecrawl subsite crawl
                current_url = seed_url
                page_no = 1
                while current_url:
                    logger.info("Harvesting page %d: %s", page_no, current_url)
                    page_links, next_page_url = crawler.extract_links(current_url, list_item_selectors=detail_link_selectors)
                    for link in page_links:
                        if link not in detail_urls:
                            detail_urls.append(link)
                    if args.max_pages > 0 and page_no >= args.max_pages:
                        break
                    if not next_page_url or next_page_url == current_url:
                        break
                    current_url = next_page_url
                    page_no += 1
    elif any(pat in args.start_url for pat in api_list_patterns):
        # Single API list URL passed directly
        from shanghai_policy_crawler.shanghai_gov_api import ShanghaiGovApiClient
        from urllib.parse import parse_qs
        parsed = urlparse(args.start_url)
        qs = parse_qs(parsed.query)
        level = qs.get("level", ["city"])[0]
        site_id_map = {"city": ["0001"], "department": None, "district": None}
        site_ids = site_id_map.get(level, ["0001"])
        logger.info("Step 1: API harvest for %s (level=%s)", args.start_url, level)
        api_client = ShanghaiGovApiClient(site_id_list=site_ids, page_size=20)
        for rec in api_client.iter_records(max_pages=args.max_pages, min_year=args.min_year):
            url = rec["detail_url"]
            if url not in detail_urls:
                detail_urls.append(url)
    else:
        if any(pat in args.start_url for pat in detail_patterns):
            logger.info("Detected start_url is a detail page: %s", args.start_url)
            detail_urls.append(args.start_url)
        else:
            logger.info("Step 1: Harvesting detail links from start_url: %s", args.start_url)
            current_url = args.start_url
            page_no = 1
            while current_url:
                logger.info("Harvesting page %d: %s", page_no, current_url)
                page_links, next_page_url = crawler.extract_links(current_url, list_item_selectors=detail_link_selectors)
                if page_links:
                    for link in page_links:
                        if link not in detail_urls:
                            detail_urls.append(link)
                
                if args.max_pages > 0 and page_no >= args.max_pages:
                    logger.info("Reached max list page limit: %d", args.max_pages)
                    break
                    
                if not next_page_url:
                    logger.info("No next page URL detected. Pagination end.")
                    break
                    
                current_url = next_page_url
                page_no += 1

    if not detail_urls:
        logger.error("No links were extracted from the start URL. Exiting.")
        sys.exit(0)

    logger.info("Found %d detail URLs to scrape. Starting phase 2.", len(detail_urls))
    
    # Setup output paths
    out_path = Path(args.output_file)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    records = []
    success_count = 0
    
    # Open target file for incremental saving
    with out_path.open("w", encoding="utf-8") as f:
        for idx, url in enumerate(detail_urls, start=1):
            logger.info("[%d/%d] Scraping details: %s", idx, len(detail_urls), url)
            record = crawler.extract_detail(url)
            if record:
                f.write(json.dumps(compact_dict(record), ensure_ascii=False) + "\n")
                f.flush()
                records.append(record)
                success_count += 1
                logger.info("Successfully saved record: '%s'", record.get("title"))
            else:
                logger.warning("Skipped/Failed record extraction: %s", url)

    logger.info("Crawling completed. Successfully extracted %d/%d records.", success_count, len(detail_urls))

    # Compile summaries (Excel, MD, CSV)
    logger.info("Step 3: Compiling structured summary workbooks...")
    build_exporters_from_records(records, Path(args.summary_prefix))
    logger.info("Post-processing exporters completed. Task success.")


if __name__ == "__main__":
    main()
